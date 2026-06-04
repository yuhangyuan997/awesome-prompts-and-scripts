#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按条件拆分Excel文件为多个文件

功能：根据Excel中某一列的值，将数据拆分为多个独立的Excel文件
用法：python3 split_excel.py [源文件] [列名] [输出目录]

参数：
  源文件    - 要拆分的Excel文件路径
  列名      - 用作拆分依据的列名（第一行为表头）
  输出目录  - 拆分文件的保存目录（默认：split_output）

示例：
  python3 split_excel.py data.xlsx 部门
  python3 split_excel.py data.xlsx 城市 ./output
"""

import os
import sys
from openpyxl import load_workbook


def split_excel(source_file, split_column, output_dir):
    """按指定列拆分Excel"""
    # 加载源文件
    try:
        wb = load_workbook(source_file, read_only=True)
    except Exception as e:
        print("错误：无法读取文件 '{}'：{}".format(source_file, e))
        sys.exit(1)

    ws = wb.active
    if ws is None:
        print("错误：文件中没有工作表")
        sys.exit(1)

    # 读取表头
    header = []
    for cell in next(ws.iter_rows(max_row=1, values_only=True)):
        header.append(cell)

    if not header:
        print("错误：文件为空或没有表头")
        sys.exit(1)

    # 查找拆分列索引
    try:
        split_idx = header.index(split_column)
    except ValueError:
        print("错误：未找到列 '{}'，可用列：{}".format(split_column, ", ".join(str(h) for h in header)))
        sys.exit(1)

    print("表头：{}".format(" | ".join(str(h) for h in header)))
    print("按列 '{}' 拆分...".format(split_column))

    # 创建输出目录
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 收集分组数据
    groups = {}
    total_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        # 跳过全空行
        if all(cell is None for cell in row):
            continue

        key = str(row[split_idx]) if row[split_idx] is not None else "空值"
        if key not in groups:
            groups[key] = []
        groups[key].append(row)

    wb.close()

    if not groups:
        print("错误：没有找到有效数据行")
        sys.exit(1)

    print("共 {} 行数据，按 {} 个分组保存...".format(total_rows, len(groups)))

    # 为每个分组创建独立Excel文件
    for idx, (group_name, rows) in enumerate(groups.items(), 1):
        # 过滤文件名中的非法字符
        safe_name = "".join(c if c.isalnum() or c in "-_. " else "_" for c in group_name)
        safe_name = safe_name.strip()[:50] or "unnamed"

        output_file = os.path.join(output_dir, "{}.xlsx".format(safe_name))

        try:
            from openpyxl import Workbook
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = "Sheet1"

            # 写入表头
            ws_out.append(header)
            # 写入数据行
            for row in rows:
                ws_out.append(row)

            wb_out.save(output_file)
            print("[{}/{}] {} -> {} 行".format(idx, len(groups), group_name, len(rows)))
        except Exception as e:
            print("  警告：保存分组 '{}' 失败：{}".format(group_name, e))

    print("\n✓ 拆分完成！共生成 {} 个文件".format(len(groups)))
    print("  输出目录：{}".format(os.path.abspath(output_dir)))


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("用法：python3 split_excel.py [源文件] [列名] [输出目录]")
        print("示例：python3 split_excel.py data.xlsx 部门")
        sys.exit(1)

    source_file = sys.argv[1]
    if not os.path.isfile(source_file):
        print("错误：文件 '{}' 不存在".format(source_file))
        sys.exit(1)

    split_column = sys.argv[2] if len(sys.argv) > 2 else None
    if not split_column:
        print("错误：请指定拆分依据的列名")
        sys.exit(1)

    output_dir = sys.argv[3] if len(sys.argv) > 3 else "split_output"

    split_excel(source_file, split_column, output_dir)


if __name__ == "__main__":
    main()
