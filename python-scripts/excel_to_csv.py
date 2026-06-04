#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel批量转CSV格式

功能：将一个或多个Excel文件（.xlsx/.xls）批量转换为CSV格式
用法：python3 excel_to_csv.py [源路径] [输出目录]

参数：
  源路径    - Excel文件或包含Excel文件的目录（默认：当前目录）
  输出目录  - CSV文件保存目录（默认：csv_output）

示例：
  python3 excel_to_csv.py data.xlsx
  python3 excel_to_csv.py ./excel_files ./csv_files
"""

import os
import sys
import glob
import csv
from openpyxl import load_workbook


def excel_to_csv(input_path, output_dir):
    """将Excel文件转换为CSV"""
    # 确定要处理的文件列表
    if os.path.isfile(input_path):
        files = [input_path]
    elif os.path.isdir(input_path):
        pattern = os.path.join(input_path, "*")
        files = glob.glob(os.path.join(input_path, "*.xlsx")) + \
                glob.glob(os.path.join(input_path, "*.xls"))
        if not files:
            print("错误：目录 '{}' 中未找到Excel文件".format(input_path))
            sys.exit(1)
    else:
        print("错误：路径 '{}' 不存在".format(input_path))
        sys.exit(1)

    # 创建输出目录
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    success_count = 0
    fail_count = 0

    for file_path in files:
        file_name = os.path.basename(file_path)
        base_name = os.path.splitext(file_name)[0]

        print("[处理] {}".format(file_name))

        try:
            wb = load_workbook(file_path, read_only=True)
        except Exception as e:
            print("  ✗ 无法读取：{}".format(e))
            fail_count += 1
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 每个Sheet单独生成一个CSV
            csv_name = "{}_{}.csv".format(base_name, sheet_name[:50])
            csv_path = os.path.join(output_dir, csv_name)

            try:
                with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    row_count = 0
                    for row in ws.iter_rows(values_only=True):
                        # 处理None值，转换为空字符串
                        cleaned = ["" if cell is None else str(cell) for cell in row]
                        writer.writerow(cleaned)
                        row_count += 1

                print("  -> {} ({} 行)".format(csv_name, row_count))
                success_count += 1
            except Exception as e:
                print("  ✗ 写入CSV失败：{}".format(e))
                fail_count += 1

        wb.close()

    print("\n✓ 转换完成！成功：{} 文件，失败：{} 文件".format(success_count, fail_count))
    print("  输出目录：{}".format(os.path.abspath(output_dir)))


def main():
    """主函数"""
    input_path = "."
    output_dir = "csv_output"

    if len(sys.argv) > 1:
        input_path = sys.argv[1]
    if len(sys.argv) > 2:
        output_dir = sys.argv[2]

    excel_to_csv(input_path, output_dir)


if __name__ == "__main__":
    main()
