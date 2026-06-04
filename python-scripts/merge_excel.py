#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
合并多个Excel文件为一个文件

功能：将指定目录下的所有Excel文件（.xlsx）合并为一个Excel文件
用法：python3 merge_excel.py [源目录] [输出文件]

参数：
  源目录    - 包含Excel文件的目录路径（默认：当前目录）
  输出文件  - 合并后的Excel文件路径（默认：merged_output.xlsx）

示例：
  python3 merge_excel.py ./data ./合并结果.xlsx
  python3 merge_excel.py                # 使用默认参数
"""

import os
import sys
import glob
from openpyxl import Workbook, load_workbook


def get_excel_files(source_dir):
    """获取指定目录下所有.xlsx文件"""
    pattern = os.path.join(source_dir, "*.xlsx")
    files = glob.glob(pattern)
    # 排除已存在的输出文件（避免循环合并）
    return [f for f in files]


def merge_excel_files(source_dir, output_file):
    """合并多个Excel文件"""
    excel_files = get_excel_files(source_dir)

    if not excel_files:
        print("错误：在目录 '{}' 中未找到任何Excel文件".format(source_dir))
        sys.exit(1)

    print("找到 {} 个Excel文件，开始合并...".format(len(excel_files)))

    # 创建新工作簿
    wb_out = Workbook()
    # 删除默认Sheet
    wb_out.remove(wb_out.active)

    total_rows = 0

    for idx, file_path in enumerate(excel_files, 1):
        file_name = os.path.basename(file_path)
        print("[{}/{}] 处理: {}".format(idx, len(excel_files), file_name))

        try:
            wb_in = load_workbook(file_path, read_only=True)
        except Exception as e:
            print("  警告：无法读取文件 '{}'：{}".format(file_name, e))
            continue

        for sheet_name in wb_in.sheetnames:
            ws_in = wb_in[sheet_name]

            # 使用原文件名+Sheet名作为新Sheet名（限制31字符）
            new_sheet_name = "{}_{}".format(
                os.path.splitext(file_name)[0], sheet_name
            )[:31]

            ws_out = wb_out.create_sheet(title=new_sheet_name)

            row_count = 0
            for row in ws_in.iter_rows(values_only=True):
                ws_out.append(row)
                row_count += 1

            total_rows += row_count
            print("  -> Sheet '{}'：{} 行数据".format(sheet_name, row_count))

        wb_in.close()

    # 保存合并结果
    try:
        wb_out.save(output_file)
        print("\n✓ 合并完成！共 {} 行数据".format(total_rows))
        print("  输出文件：{}".format(os.path.abspath(output_file)))
    except Exception as e:
        print("错误：无法保存输出文件：{}".format(e))
        sys.exit(1)


def main():
    """主函数：解析参数并执行合并"""
    # 解析命令行参数
    source_dir = "."
    output_file = "merged_output.xlsx"

    if len(sys.argv) > 1:
        source_dir = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]

    # 检查源目录
    if not os.path.isdir(source_dir):
        print("错误：目录 '{}' 不存在或不是一个目录".format(source_dir))
        sys.exit(1)

    # 确保输出文件路径完整
    output_file = os.path.abspath(output_file)
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)

    merge_excel_files(source_dir, output_file)


if __name__ == "__main__":
    main()
